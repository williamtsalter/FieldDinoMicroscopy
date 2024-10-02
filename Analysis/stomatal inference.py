import os
import cv2
import numpy as np
import pandas as pd
from ultralytics import YOLO
import torch
import openpyxl as xl
import math

# Replace with your paths and model name
model_path = "runs/segment/20240310_xlarge_239epochs/weights/best.pt"  # Update with your YOLOv8 model (.pt)
image_dir = "images"
output_dir = "inference/output/labels"
ellipses_dir = "inference/output/ellipses"
allresults_excel_file_path = "inference/output/allresults.xlsx"
summaryresults_excel_file_path = "inference/output/summaryresults.xlsx"

#Replace with scale of images (1 pixel = 0.677953691695135 µm)
image_scale = 0.677953691695135

# Calculate true area of image in mm^2 for density calculation
image_area = (2592 * 1944) * (image_scale / 1000) ** 2

# Constants for Gsmax calculation
# Pi
PI = math.pi
# Diffusivity of water in air (d in m^2 s^-1 at 25°C) = 0.0000249
d = 0.0000249
# Molar volume of air (v in m^3 mol^-1 at 25°C) = 0.0244
v = 0.0244
# Guard cell length to pore length scaling factor (p) = 0.543 (R2 = 0.9817)
p = 0.543

# Create output directories if it doesn't exist
os.makedirs(output_dir, exist_ok=True)
os.makedirs(ellipses_dir, exist_ok=True)

# Initialize necessary objects
model = YOLO(model_path)  # Adapt loading based on your model structure

# Define column headings
all_column_headings = ["Filename", "Stomatal Count", "Stomatal Area (µm^2)", "Stomatal Width (µm)", "Stomatal Length (µm)", "Pore length (µm)", "Pore width (µm)", "Pore area (µm^2)", "Pore depth (µm)"]
summary_column_headings = ["Filename", "Stomatal Count", "Stomatal Density (per mm^2)", "Mean Stomatal Area (µm^2)", "Stdev Stomatal Area (µm^2)", "Mean Stomatal Width (µm)", "Stdev Stomatal Width (µm)", "Mean Stomatal Length (µm)", "Stdev Stomatal Length (µm)", "Mean pore area (µm^2)", "Stdev pore area (µm^2)", "gsmax (mol m^-2 s^-1)"]

# Check if Excel files exist and update row_count
# All data
if os.path.exists(allresults_excel_file_path):
    source_file_all = xl.load_workbook(allresults_excel_file_path)
    sheet_name_all = "Sheet1"
    sheet_all = source_file_all[sheet_name_all]
    row_count_all = sheet_all.max_row
    source_file_all.close()
else:
    source_file_all = xl.Workbook()
    sheet_name_all = "Sheet1"
    sheet_all = source_file_all.active
    sheet_all.title = sheet_name_all
    sheet_all.append(all_column_headings)
    row_count_all = 1
    source_file_all.save(allresults_excel_file_path)
    source_file_all.close()

# Summary data
if os.path.exists(summaryresults_excel_file_path):
    source_file_summary = xl.load_workbook(summaryresults_excel_file_path)
    sheet_name_summary = "Sheet1"
    sheet_summary = source_file_summary[sheet_name_summary]
    row_count_summary = sheet_summary.max_row
    source_file_summary.close()
else:
    source_file_summary = xl.Workbook()
    sheet_name_summary = "Sheet1"
    sheet_summary = source_file_summary.active
    sheet_summary.title = sheet_name_summary
    sheet_summary.append(summary_column_headings)
    row_count_summary = 1
    source_file_summary.save(summaryresults_excel_file_path)
    source_file_summary.close()

# Iterate through images in the directory
for image_filename in sorted(os.listdir(image_dir)):
    if image_filename.endswith((".jpg", ".jpeg", ".png")):

        # Clear results lists before new image
        allresults = []
        summaryresults = []

        # Save image filename as image_path
        image_path = os.path.join(image_dir, image_filename)
        image_save_path = os.path.join(output_dir, image_filename)
        ellipses_path = os.path.join(ellipses_dir, image_filename)

        # Open and preprocess image
        image = cv2.imread(image_path)
        image = cv2.cvtColor(image, cv2.COLOR_BGR2RGB)

        # Perform inference with YOLOv8
        results_data = model.predict(image_path,imgsz = (1952,2592), conf = 0.4, retina_masks = True, device='cuda:0',save = False)
        save_results = results_data[0].plot(font_size=20, pil=True)
        cv2.imwrite(image_save_path, save_results)

        # Count detected objects and calculate stomatal density in pores/mm^2
        stomatal_count = 0
        prediction_count = 0
        stomatal_density = 0
        stomatal_area = 0
        stomatal_length = 0
        stomatal_width = 0
        pore_length = 0
        pore_length_in_m = 0
        pore_width = 0
        pore_width_in_m = 0
        pore_area = 0
        pore_area_in_m2 = 0
        pore_depth = 0
        pore_depth_in_m = 0
        total_pore_area = 0
        total_pore_depth = 0
        amax = 0
        l = 0


        # Run through every detection to measure size of each pore
        for prediction in results_data:
            # Handle non-rectangular masks
            label = prediction.names[prediction.boxes.cls.tolist().pop()]

            # Iterate over each mask and extract information
            for mask in prediction.masks.xy:
                prediction_count += 1
                # Convert mask to numpy array
                mask_np = np.array(mask).astype(np.int32)

                # Create binary mask image
                mask_img = np.zeros_like(image)
                cv2.fillPoly(mask_img, [mask_np], (255, 255, 255))
                # Crop slightly to accurately remove edge detections
                cropped_mask_img = mask_img[10:1942, 10:2582]
                # cv2.imshow("Ellipse", cropped_mask_img)
                # cv2.waitKey(0)
                # cv2.destroyAllWindows()

                # Calculate contour
                contours, _ = cv2.findContours(cropped_mask_img[:, :, 0], cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)

                filtered_contours = []
                for contour in contours:
                    # Check if any pixel in the mask image lies on or beyond the edges of the image
                    outside_image = (cropped_mask_img[0, :].any() | cropped_mask_img[-1, :].any() | cropped_mask_img[:, 0].any() | cropped_mask_img[:,-1].any())

                    if outside_image:
                        print("OUTSIDE IMAGE")

                    if not outside_image:
                        filtered_contours.append(contour)

                # Calculate area using contour approximation in µm^2
                if filtered_contours:
                    # Calculate area using contour approximation in µm^2
                    stomatal_area = cv2.contourArea(filtered_contours[0]) * (image_scale ** 2)  # Assuming single contour per object

                # Use the first contour to calculate the area
                    contour_points = len(filtered_contours[0])
                    if contour_points > 10:
                        new_stomata = 1
                        ellipse = cv2.fitEllipse(filtered_contours[0])
                        (x, y), (width, height), angle = ellipse
                        cropped_img = image[10:1942, 10:2582]
                        cv2.ellipse(cropped_img, ellipse, (0, 0, 255), 3)
                        # cv2.imshow("Ellipse", image)
                        # cv2.waitKey(0)
                        # cv2.destroyAllWindows()

                        # Scale measurements to µm based upon image scale
                        stomatal_width = width * image_scale
                        stomatal_length = height * image_scale

                        # Estimate pore traits
                        pore_length = stomatal_length * p
                        pore_length_in_m = pore_length / 1000000
                        pore_width = pore_length / 2
                        pore_width_in_m = pore_width / 1000000
                        pore_area = PI * (pore_length / 2) ** 2
                        pore_area_in_m2 = PI * (pore_length_in_m / 2) ** 2
                        pore_depth = pore_width / 2
                        pore_depth_in_m = pore_width_in_m / 2

                        # Update total pore area (amax) for leaf
                        stomatal_count += new_stomata
                        total_pore_area += pore_area_in_m2
                        total_pore_depth += pore_depth_in_m

                        # Append allresults to list
                        allresults.append({
                            "filename": image_filename,
                            "stomatal_count": stomatal_count,
                            "stomatal_area_(µm^2)": stomatal_area,
                            "stomatal_width_(µm)": stomatal_width,
                            "stomatal_length_(µm)": stomatal_length,
                            "pore_length_(µm)": pore_length,
                            "pore_width_(µm)": pore_width,
                            "pore_area_(µm^2)": pore_area,
                            "pore_depth_(µm)": pore_depth
                        })

        # Save ellipses
        # cv2.imshow("Image", image)
        # cv2.waitKey(0)
        cv2.imwrite(ellipses_path,cropped_img)

        # Calculate summary stats for each image
        df_allresults = pd.DataFrame(allresults)
        stomatal_area_mean = df_allresults["stomatal_area_(µm^2)"].mean()
        stomatal_area_stdev = df_allresults["stomatal_area_(µm^2)"].std()
        stomatal_width_mean = df_allresults["stomatal_width_(µm)"].mean()
        stomatal_width_stdev = df_allresults["stomatal_width_(µm)"].std()
        stomatal_length_mean = df_allresults["stomatal_length_(µm)"].mean()
        stomatal_length_stdev = df_allresults["stomatal_length_(µm)"].std()
        pore_area_mean = df_allresults["pore_area_(µm^2)"].mean()
        pore_area_stdev = df_allresults["pore_area_(µm^2)"].std()

        #Stomatal density calculation
        stomatal_density = stomatal_count / image_area
        stomatal_density_per_m2 = stomatal_density * 1000000

        # Calculate Gsmax
        # Maximum stomatal pore area (amax)
        amax = total_pore_area / stomatal_count
        l = total_pore_depth / stomatal_count
        gsmax_numerator = (d / v) * stomatal_density_per_m2 * amax
        gsmax_denominator = l + (PI / 2) * np.sqrt(amax/PI)
        gsmax_in_mol = gsmax_numerator / gsmax_denominator

        print("Stomatal count:", stomatal_count)
        print("Gsmax:", gsmax_in_mol, " mol m^-2 s^-1")


        # Append to list
        summaryresults.append({
            "filename": image_filename,
            "stomatal_count": stomatal_count,
            "stomatal_density_(pores/mm^2)": stomatal_density,
            "mean_stomatal_area_(µm^2)": stomatal_area_mean,
            "stdev_stomatal_area_(µm^2)": stomatal_area_stdev,
            "mean_stomatal_width_(µm)": stomatal_width_mean,
            "stdev_stomatal_width_(µm)": stomatal_width_stdev,
            "mean_stomatal_length_(µm)": stomatal_length_mean,
            "stdev_stomatal_length_(µm)": stomatal_length_stdev,
            "mean_pore_area": pore_area_mean,
            "stdev_pore_area": pore_area_stdev,
            "gsmax_mmol": gsmax_in_mol
        })

        # Append all results to Excel file
        with pd.ExcelWriter(allresults_excel_file_path, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
            pd.DataFrame(allresults).to_excel(writer, sheet_name=sheet_name_all, header=None, index=False, startrow=row_count_all)

        # Append summary results to Excel file
        with pd.ExcelWriter(summaryresults_excel_file_path, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
            pd.DataFrame(summaryresults).to_excel(writer, sheet_name=sheet_name_summary, header=None, index=False, startrow=row_count_summary)

        # Update row counts
        row_count_all += stomatal_count
        row_count_summary += 1

        print(f"Saved results to Excel files   :D")

        torch.cuda.empty_cache()
